"""
Save Template Operator for EM-blender-tools

Saves the standard stratigraphy and site properties Excel templates to a user-chosen directory.
Uses the templates bundled with s3dgraphy or generates them dynamically.
"""

import bpy  # type: ignore
import os
import shutil


class EMTOOLS_OT_save_stratigraphy_template(bpy.types.Operator):
    """Save the standard stratigraphy Excel template to disk"""
    bl_idname = "emtools.save_stratigraphy_template"
    bl_label = "Save Stratigraphy Template"
    bl_description = (
        "Save the standard 24-column stratigraphy Excel template "
        "(template_stratigraphy.xlsx) to a chosen directory"
    )
    bl_options = {'REGISTER'}

    directory: bpy.props.StringProperty(
        subtype='DIR_PATH',
        description="Directory to save the template"
    )  # type: ignore

    def invoke(self, context, event):
        context.window_manager.fileselect_add(self)
        return {'RUNNING_MODAL'}

    def execute(self, context):
        if not self.directory:
            self.report({'ERROR'}, "No directory selected")
            return {'CANCELLED'}

        try:
            template_path = self._find_template("template_stratigraphy.xlsx")
            if template_path:
                dest = os.path.join(self.directory, "template_stratigraphy.xlsx")
                shutil.copy2(template_path, dest)
                self.report({'INFO'}, f"Template saved to: {dest}")
                return {'FINISHED'}
            else:
                # Generate template dynamically
                return self._generate_template(self.directory, "stratigraphy")
        except Exception as e:
            self.report({'ERROR'}, f"Failed to save template: {str(e)}")
            return {'CANCELLED'}

    def _find_template(self, filename):
        """Try to find the template file in s3dgraphy package or common locations."""
        # Try s3dgraphy package location
        try:
            import s3dgraphy
            pkg_dir = os.path.dirname(os.path.dirname(s3dgraphy.__file__))
            # Check in package root (where pip installs)
            candidate = os.path.join(pkg_dir, filename)
            if os.path.exists(candidate):
                return candidate
            # Check in s3dgraphy source root
            src_root = os.path.dirname(pkg_dir)
            candidate = os.path.join(src_root, filename)
            if os.path.exists(candidate):
                return candidate
        except ImportError:
            pass
        return None

    def _generate_template(self, directory, template_type):
        """Generate template dynamically using openpyxl."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font
        except ImportError:
            self.report({'ERROR'}, "openpyxl not available. Cannot generate template.")
            return {'CANCELLED'}

        wb = Workbook()
        ws = wb.active
        ws.title = "Stratigraphy"

        headers = [
            "ID", "TYPE", "DESCRIPTION", "PERIOD", "PERIOD_START", "PERIOD_END",
            "PHASE", "PHASE_START", "PHASE_END", "SUBPHASE", "SUBPHASE_START",
            "SUBPHASE_END", "OVERLIES", "OVERLAIN_BY", "CUTS", "CUT_BY",
            "FILLS", "FILLED_BY", "ABUTS", "ABUTTED_BY", "BONDED_TO",
            "EQUALS", "EXTRACTOR", "DOCUMENT"
        ]

        bold = Font(bold=True)
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = bold

        dest = os.path.join(directory, "template_stratigraphy.xlsx")
        wb.save(dest)
        self.report({'INFO'}, f"Template generated and saved to: {dest}")
        return {'FINISHED'}


class EMTOOLS_OT_save_em_paradata_template(bpy.types.Operator):
    """Save the em_paradata Excel template to disk"""
    bl_idname = "emtools.save_em_paradata_template"
    bl_label = "Save Paradata Template"
    bl_description = (
        "Save the standard em_paradata Excel template "
        "(template_em_paradata.xlsx) to a chosen directory"
    )
    bl_options = {'REGISTER'}

    directory: bpy.props.StringProperty(
        subtype='DIR_PATH',
        description="Directory to save the template"
    )  # type: ignore

    def invoke(self, context, event):
        context.window_manager.fileselect_add(self)
        return {'RUNNING_MODAL'}

    def execute(self, context):
        if not self.directory:
            self.report({'ERROR'}, "No directory selected")
            return {'CANCELLED'}

        try:
            template_path = self._find_template("template_em_paradata.xlsx")
            if template_path:
                dest = os.path.join(self.directory, "template_em_paradata.xlsx")
                shutil.copy2(template_path, dest)
                self.report({'INFO'}, f"Template saved to: {dest}")
                return {'FINISHED'}
            else:
                return self._generate_template(self.directory)
        except Exception as e:
            self.report({'ERROR'}, f"Failed to save template: {str(e)}")
            return {'CANCELLED'}

    def _find_template(self, filename):
        """Try to find the template file in s3dgraphy package."""
        try:
            import s3dgraphy
            pkg_dir = os.path.dirname(os.path.dirname(s3dgraphy.__file__))
            candidate = os.path.join(pkg_dir, filename)
            if os.path.exists(candidate):
                return candidate
            src_root = os.path.dirname(pkg_dir)
            candidate = os.path.join(src_root, filename)
            if os.path.exists(candidate):
                return candidate
        except ImportError:
            pass
        return None

    def _generate_template(self, directory):
        """Generate template dynamically using openpyxl."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font
        except ImportError:
            self.report({'ERROR'}, "openpyxl not available. Cannot generate template.")
            return {'CANCELLED'}

        wb = Workbook()
        ws = wb.active
        ws.title = "Paradata"

        headers = [
            "US_ID", "PROPERTY_TYPE", "VALUE", "COMBINER_REASONING",
            "EXTRACTOR_1", "DOCUMENT_1", "EXTRACTOR_2", "DOCUMENT_2"
        ]

        bold = Font(bold=True)
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = bold

        dest = os.path.join(directory, "template_em_paradata.xlsx")
        wb.save(dest)
        self.report({'INFO'}, f"Template generated and saved to: {dest}")
        return {'FINISHED'}


class EMTOOLS_OT_save_em_data_template(bpy.types.Operator):
    """Save the unified em_data.xlsx template to disk"""
    bl_idname = "emtools.save_em_data_template"
    bl_label = "Save em_data.xlsx Template"
    bl_description = (
        "Save the StratiMiner unified Excel template "
        "(em_data_template.xlsx — 5 typed sheets: Units, Epochs, Claims, "
        "Authors, Documents) to a chosen directory. Fill it by hand or "
        "use it as reference for the AI-extracted output"
    )
    bl_options = {'REGISTER'}

    directory: bpy.props.StringProperty(
        subtype='DIR_PATH',
        description="Directory to save the template"
    )  # type: ignore

    def invoke(self, context, event):
        context.window_manager.fileselect_add(self)
        return {'RUNNING_MODAL'}

    def execute(self, context):
        if not self.directory:
            self.report({'ERROR'}, "No directory selected")
            return {'CANCELLED'}

        # The unified template lives inside s3dgraphy at
        # s3dgraphy/templates/em_data_template.xlsx (shipped with the
        # library, not regenerated here).
        try:
            import s3dgraphy
            pkg_dir = os.path.dirname(s3dgraphy.__file__)
            src = os.path.join(pkg_dir, "templates", "em_data_template.xlsx")
            if not os.path.exists(src):
                self.report({'ERROR'},
                            f"Template not found at {src}. "
                            "Ensure s3dgraphy is installed correctly.")
                return {'CANCELLED'}
            dest = os.path.join(self.directory, "em_data_template.xlsx")
            shutil.copy2(src, dest)
            self.report({'INFO'}, f"Template saved to: {dest}")
            return {'FINISHED'}
        except Exception as e:
            self.report({'ERROR'}, f"Failed to save template: {e}")
            return {'CANCELLED'}


classes = (
    EMTOOLS_OT_save_stratigraphy_template,
    EMTOOLS_OT_save_em_paradata_template,
    EMTOOLS_OT_save_em_data_template,
)


def register():
    for cls in classes:
        bpy.utils.register_class(cls)


def unregister():
    for cls in reversed(classes):
        bpy.utils.unregister_class(cls)
